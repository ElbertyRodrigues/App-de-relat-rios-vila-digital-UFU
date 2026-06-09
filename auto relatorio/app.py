"""
Vila Digital UFU — Gerador de Relatório
v10.0 — turno (Manhã/Tarde/Noite) no relatório de disciplinas
"""

import sys, os, time, re, threading
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
from pathlib import Path

def instalar_dependencias():
    import subprocess
    for pkg in ["selenium", "openpyxl", "webdriver-manager"]:
        try:
            __import__(pkg.replace("-", "_"))
        except ImportError:
            subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])

instalar_dependencias()

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

CAMPI = {
    "Santa Mônica (Bloco 3Q)": {
        "cidade": "Uberlândia",
        "salas": {
            "3Q103B": {"url": "https://proplad.ufu.br/vila-digital/laboratorio/campus-santa-monica-bloco-3q/3q103b", "computadores": 18},
            "3Q104":  {"url": "https://proplad.ufu.br/vila-digital/laboratorio/campus-santa-monica-bloco-3q/3q104",  "computadores": 25},
            "3Q105":  {"url": "https://proplad.ufu.br/vila-digital/laboratorio/campus-santa-monica-bloco-3q/3q105",  "computadores": 25},
            "3Q106":  {"url": "https://proplad.ufu.br/vila-digital/laboratorio/campus-santa-monica-bloco-3q/3q106",  "computadores": 25},
        }
    },
    "Educação Física": {
        "cidade": "Uberlândia",
        "salas": {
            "1N01": {"url": "https://proplad.ufu.br/vila-digital/laboratorio/educacao-fisica/1n01", "computadores": 18},
        }
    },
    "Umuarama (Bloco 8C)": {
        "cidade": "Uberlândia",
        "salas": {
            "8C01": {"url": "https://proplad.ufu.br/vila-digital/laboratorio/umuarama/8c01",                 "computadores": 20},
            "8C02": {"url": "https://proplad.ufu.br/vila-digital/laboratorio/campus-umuarama-bloco-8c/8c02", "computadores": 20},
            "8C03": {"url": "https://proplad.ufu.br/vila-digital/laboratorio/umuarama/8c03",                 "computadores": 35},
        }
    },
}

MESES_PT = {
    "janeiro":1,"fevereiro":2,"março":3,"abril":4,
    "maio":5,"junho":6,"julho":7,"agosto":8,
    "setembro":9,"outubro":10,"novembro":11,"dezembro":12
}


def _hora_para_turno(hora_str):
    """Converte 'HH:MM' em turno: Manhã (06-11h), Tarde (12-17h), Noite (18-23h)."""
    try:
        h = int(hora_str.split(":")[0])
        if 6 <= h < 12:
            return "Manhã"
        elif 12 <= h < 18:
            return "Tarde"
        else:
            return "Noite"
    except Exception:
        return ""

def _calcular_turno(hora_inicio, hora_fim):
    """Retorna turno único ou combinado (ex: 'Manhã/Tarde') se cruzar dois turnos."""
    t_ini = _hora_para_turno(hora_inicio)
    t_fim = _hora_para_turno(hora_fim)
    if not t_ini:
        return ""
    if t_fim and t_fim != t_ini:
        return f"{t_ini}/{t_fim}"
    return t_ini

def fmt_tempo(s):
    s = int(s)
    return f"{s}s" if s < 60 else f"{s//60}m {s%60:02d}s"

def criar_driver(headless):
    opts = Options()
    if headless:
        opts.add_argument("--headless=new")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1400,900")
    opts.add_argument("--log-level=3")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()), options=opts)
    
    driver.get("about:blank")
    driver.execute_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    return driver

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Vila Digital UFU — Gerador de Relatório")
        self.resizable(False, True)
        self.configure(bg="#1F3864")
        self._total_steps   = 0
        self._current_step  = 0
        self._tempo_inicio  = None
        self._timer_running = False
        self.modo_var = tk.StringVar(value="tudo")
        self._build_ui()
        self._center()

    def _center(self):
        self.update_idletasks()
        w, h = 520, 820
        self.geometry(f"{w}x{h}+{(self.winfo_screenwidth()-w)//2}+{(self.winfo_screenheight()-h)//2}")

    def _build_ui(self):
        tk.Label(self, text="Vila Digital UFU", bg="#1F3864",
                 fg="white", font=("Arial",18,"bold")).pack(pady=(20,2))
        tk.Label(self, text="Gerador de Relatório de Utilização",
                 bg="#1F3864", fg="#A8C8E8", font=("Arial",10)).pack(pady=(0,14))

        card = tk.Frame(self, bg="white"); card.pack(padx=30, fill="x")
        inner = tk.Frame(card, bg="white"); inner.pack(padx=25, pady=20, fill="x")

        tk.Label(inner, text="Campus", bg="white", fg="#1F3864",
                 font=("Arial",11,"bold")).grid(row=0, column=0, columnspan=4, sticky="w", pady=(0,6))
        self.campus_var = tk.StringVar(value=list(CAMPI.keys())[0])
        cb = ttk.Combobox(inner, textvariable=self.campus_var,
                          values=list(CAMPI.keys()), state="readonly", width=38)
        cb.grid(row=1, column=0, columnspan=4, sticky="ew", pady=(0,4))
        cb.bind("<<ComboboxSelected>>", self._on_campus_change)
        self.lbl_cidade = tk.Label(inner, text="", bg="white", fg="#7F7F7F", font=("Arial",9,"italic"))
        self.lbl_cidade.grid(row=2, column=0, columnspan=4, sticky="w", pady=(0,2))
        self.lbl_aviso = tk.Label(inner, text="", bg="#FFF8E1", fg="#795548",
                                  font=("Arial",8,"italic"), wraplength=400, justify="left", padx=6, pady=4)
        self.lbl_aviso.grid(row=3, column=0, columnspan=4, sticky="ew", pady=(0,4))

        ttk.Separator(inner, orient="horizontal").grid(row=4, column=0, columnspan=4, sticky="ew", pady=(8,8))
        tk.Label(inner, text="Período do relatório", bg="white", fg="#1F3864",
                 font=("Arial",11,"bold")).grid(row=5, column=0, columnspan=4, sticky="w", pady=(0,8))

        tk.Label(inner, text="Data inicial:", bg="white", font=("Arial",10)).grid(row=6, column=0, sticky="w")
        self.ini_d = ttk.Combobox(inner, width=4, values=[f"{i:02d}" for i in range(1,32)]); self.ini_d.set("01")
        self.ini_d.grid(row=6, column=1, padx=(5,2))
        self.ini_m = ttk.Combobox(inner, width=4, values=[f"{i:02d}" for i in range(1,13)]); self.ini_m.set("04")
        self.ini_m.grid(row=6, column=2, padx=2)
        self.ini_a = ttk.Combobox(inner, width=6, values=["2024","2025","2026","2027"]); self.ini_a.set("2026")
        self.ini_a.grid(row=6, column=3)

        tk.Label(inner, text="Data final:", bg="white", font=("Arial",10)).grid(row=7, column=0, sticky="w", pady=(6,0))
        self.fim_d = ttk.Combobox(inner, width=4, values=[f"{i:02d}" for i in range(1,32)]); self.fim_d.set("30")
        self.fim_d.grid(row=7, column=1, padx=(5,2), pady=(6,0))
        self.fim_m = ttk.Combobox(inner, width=4, values=[f"{i:02d}" for i in range(1,13)]); self.fim_m.set("09")
        self.fim_m.grid(row=7, column=2, padx=2, pady=(6,0))
        self.fim_a = ttk.Combobox(inner, width=6, values=["2024","2025","2026","2027"]); self.fim_a.set("2026")
        self.fim_a.grid(row=7, column=3, pady=(6,0))

        ttk.Separator(inner, orient="horizontal").grid(row=8, column=0, columnspan=4, sticky="ew", pady=(12,8))
        tk.Label(inner, text="Salas a incluir", bg="white", fg="#1F3864",
                 font=("Arial",11,"bold")).grid(row=9, column=0, columnspan=4, sticky="w", pady=(0,6))
        self.salas_frame = tk.Frame(inner, bg="white")
        self.salas_frame.grid(row=10, column=0, columnspan=4, sticky="ew")
        self.vars_salas = {}

        ttk.Separator(inner, orient="horizontal").grid(row=11, column=0, columnspan=4, sticky="ew", pady=(12,8))
        tk.Label(inner, text="Modo de coleta", bg="white", fg="#1F3864",
                 font=("Arial",11,"bold")).grid(row=12, column=0, columnspan=4, sticky="w", pady=(0,6))
        modos = [
            ("📋  Só Disciplinas  (Código + Nome — arquivo separado)",  "disciplinas"),
            ("📊  Só Totais  (Relatório principal de aulas e alunos)",   "totais"),
            ("🗂️  Tudo  (ambos os arquivos)",                            "tudo"),
        ]
        for i, (texto, valor) in enumerate(modos):
            tk.Radiobutton(inner, text=texto, variable=self.modo_var, value=valor,
                           bg="white", font=("Arial",9), activebackground="white",
                           anchor="w").grid(row=13+i, column=0, columnspan=4, sticky="w")

        ttk.Separator(inner, orient="horizontal").grid(row=16, column=0, columnspan=4, sticky="ew", pady=(12,8))
        tk.Label(inner, text="Salvar em:", bg="white", fg="#1F3864",
                 font=("Arial",11,"bold")).grid(row=17, column=0, columnspan=4, sticky="w", pady=(0,4))
        df = tk.Frame(inner, bg="white"); df.grid(row=18, column=0, columnspan=4, sticky="ew")
        self.dest_var = tk.StringVar(value=str(Path.home()/"Desktop"))
        tk.Entry(df, textvariable=self.dest_var, width=34, font=("Arial",9)).pack(side="left")
        tk.Button(df, text="...", command=self._escolher_pasta, font=("Arial",9), padx=6).pack(side="left", padx=(6,0))

        self.headless_var = tk.BooleanVar(value=True)
        tk.Checkbutton(inner, text="Modo silencioso (não abrir o Chrome visível)",
                       variable=self.headless_var, bg="white", font=("Arial",9),
                       activebackground="white").grid(row=19, column=0, columnspan=4, sticky="w", pady=(10,0))

        self.btn = tk.Button(self, text="▶  Gerar Relatório", command=self._iniciar,
                             bg="#2E75B6", fg="white", font=("Arial",13,"bold"),
                             relief="flat", padx=20, pady=10,
                             activebackground="#1F3864", activeforeground="white", cursor="hand2")
        self.btn.pack(pady=(16,10), padx=30, fill="x")

        pf = tk.Frame(self, bg="#1F3864"); pf.pack(padx=30, fill="x", pady=(0,4))
        self.lbl_status = tk.Label(pf, text="", bg="#1F3864", fg="#A8C8E8", font=("Arial",9), anchor="w")
        self.lbl_status.pack(fill="x")
        style = ttk.Style(); style.theme_use("default")
        style.configure("Vila.Horizontal.TProgressbar", troughcolor="#0D1B2A", background="#00C853", thickness=20)
        self.progressbar = ttk.Progressbar(pf, style="Vila.Horizontal.TProgressbar",
                                           orient="horizontal", mode="determinate")
        self.progressbar.pack(fill="x", pady=(4,2))
        info = tk.Frame(pf, bg="#1F3864"); info.pack(fill="x", pady=(2,4))
        self.lbl_pct = tk.Label(info, text="", bg="#1F3864", fg="#00C853",
                                font=("Consolas",9,"bold"), anchor="w")
        self.lbl_pct.pack(side="left")
        self.lbl_timer = tk.Label(info, text="", bg="#1F3864", fg="#FFD54F",
                                  font=("Consolas",9), anchor="e")
        self.lbl_timer.pack(side="right")

        self.log = tk.Text(self, height=6, bg="#0D1B2A", fg="#00FF99",
                           font=("Consolas",8), state="disabled", relief="flat", padx=8, pady=6)
        self.log.pack(padx=30, pady=(0,20), fill="both", expand=True)

        self._on_campus_change()

    def _rebuild_salas(self, campus):
        for w in self.salas_frame.winfo_children(): w.destroy()
        self.vars_salas.clear()
        for i, (sala, cfg) in enumerate(CAMPI[campus]["salas"].items()):
            v = tk.BooleanVar(value=True); self.vars_salas[sala] = v
            tk.Checkbutton(self.salas_frame, text=f"{sala}  ({cfg['computadores']} PCs)",
                           variable=v, bg="white", font=("Arial",10),
                           activebackground="white").grid(row=i//2, column=(i%2)*2, columnspan=2, sticky="w")

    def _on_campus_change(self, event=None):
        campus = self.campus_var.get()
        self._rebuild_salas(campus)
        self.lbl_cidade.configure(text=f"📍 {CAMPI[campus]['cidade']}")
        self.lbl_aviso.configure(
            text="⚠️  Este campus não registra quantidade de alunos nas reservas."
            if campus == "Umuarama (Bloco 8C)" else "")

    def _escolher_pasta(self):
        p = filedialog.askdirectory(initialdir=self.dest_var.get())
        if p: self.dest_var.set(p)

    def _log(self, msg):
        self.log.configure(state="normal")
        self.log.insert("end", msg+"\n")
        self.log.see("end")
        self.log.configure(state="disabled")
        self.update()

    def _set_progresso(self, atual, total, status=""):
        pct = int(atual/total*100) if total > 0 else 0
        self.progressbar["maximum"] = total
        self.progressbar["value"]   = atual
        self.lbl_pct.configure(text=f"{pct}%  ({atual}/{total} reservas)")
        if status: self.lbl_status.configure(text=f"⏳ {status}")
        self.update()

    def _iniciar_timer(self):
        self._tempo_inicio = time.time(); self._timer_running = True; self._tick_timer()

    def _parar_timer(self):
        self._timer_running = False

    def _tick_timer(self):
        if not self._timer_running: return
        dec = time.time()-self._tempo_inicio
        txt = f"⏱ {fmt_tempo(dec)} decorrido"
        if 0 < self._current_step < self._total_steps:
            txt += f"  |  ~{fmt_tempo((dec/self._current_step)*(self._total_steps-self._current_step))} restante"
        self.lbl_timer.configure(text=txt)
        self.after(1000, self._tick_timer)

    def _iniciar(self):
        try:
            ini = datetime.strptime(f"{self.ini_d.get()}/{self.ini_m.get()}/{self.ini_a.get()}", "%d/%m/%Y")
            fim = datetime.strptime(f"{self.fim_d.get()}/{self.fim_m.get()}/{self.fim_a.get()}", "%d/%m/%Y")
        except ValueError:
            messagebox.showerror("Erro","Data inválida."); return
        if ini > fim:
            messagebox.showerror("Erro","Data inicial maior que a final."); return
        campus    = self.campus_var.get()
        salas_cfg = CAMPI[campus]["salas"]
        salas_sel = {s: salas_cfg[s] for s,v in self.vars_salas.items() if v.get()}
        if not salas_sel:
            messagebox.showerror("Erro","Selecione ao menos uma sala."); return

        self.btn.configure(state="disabled", text="⏳  Coletando dados...")
        self.progressbar["value"]=0
        self.lbl_pct.configure(text="0%  (mapeando reservas...)")
        self.lbl_status.configure(text=""); self.lbl_timer.configure(text="")
        self._current_step=0; self._total_steps=0
        self._iniciar_timer()

        threading.Thread(target=self._coletar,
                         args=(ini,fim,campus,salas_sel,self.headless_var.get(),self.modo_var.get()),
                         daemon=True).start()

    def _coletar(self, ini, fim, campus, salas, headless, modo):
        try:
            todos = coletar_dados(ini, fim, salas, headless,
                                  log_fn=self._log,
                                  progresso_fn=self._atualizar_progresso,
                                  set_total_fn=self._set_total)
            self._parar_timer()
            if not todos:
                self.after(0, lambda: messagebox.showwarning("Aviso","Nenhuma reserva encontrada no período.")); return

            slug  = campus.replace(" ","_").replace("(","").replace(")","").replace("/","")
            pasta = self.dest_var.get()
            arquivos_gerados = []

            if modo in ("totais", "tudo"):
                nome_rel = f"Relatorio_VD_{slug}_{ini.strftime('%d%m%Y')}_a_{fim.strftime('%d%m%Y')}.xlsx"
                destino_rel = os.path.join(pasta, nome_rel)
                gerar_excel(todos, salas, campus, ini, fim, destino_rel)
                arquivos_gerados.append(f"📊 {destino_rel}")
                self._log(f"   📊 Relatório principal salvo.")

            if modo in ("disciplinas", "tudo"):
                nome_disc = f"Disciplinas_VD_{slug}_{ini.strftime('%d%m%Y')}_a_{fim.strftime('%d%m%Y')}.xlsx"
                destino_disc = os.path.join(pasta, nome_disc)
                gerar_excel_disciplinas(todos, campus, ini, fim, destino_disc)
                arquivos_gerados.append(f"📋 {destino_disc}")
                self._log(f"   📋 Relatório de disciplinas salvo.")

            dec = fmt_tempo(time.time()-self._tempo_inicio)
            self._log(f"\n✅ Concluído! Tempo total: {dec}")
            self.lbl_status.configure(text=f"✅ Concluído em {dec}!")
            self.lbl_pct.configure(text=f"100%  ({self._total_steps}/{self._total_steps} reservas)")
            self.progressbar["value"] = self._total_steps
            msg_arquivos = "\n\n".join(arquivos_gerados)
            self.after(0, lambda: messagebox.showinfo("Concluído!",
                f"Arquivo(s) gerado(s):\n\n{msg_arquivos}\n\nTempo total: {dec}"))
        except Exception as e:
            self._parar_timer()
            self._log(f"\n❌ Erro: {e}")
            self.after(0, lambda: messagebox.showerror("Erro", str(e)))
        finally:
            self.after(0, lambda: self.btn.configure(state="normal", text="▶  Gerar Relatório"))

    def _set_total(self, total):
        self._total_steps = total
        self.after(0, lambda: self._set_progresso(0, total, "Coletando detalhes..."))

    def _atualizar_progresso(self, incremento=1, status=""):
        self._current_step += incremento
        c = self._current_step
        self.after(0, lambda: self._set_progresso(c, self._total_steps, status))



def slug_base(href):
    """Remove sufixos de versão de 1 dígito (-0 a -9) do slug da URL.
    Ex: evento-egen-2025-0 -> evento-egen-2025
        famat31304matematica-0 -> famat31304matematica
    Não remove anos ou códigos longos (-2025, -3305).
    """
    s = href.rstrip("/").split("/")[-1]
    prev = None
    while prev != s:
        prev = s
        s = re.sub(r'-[0-9]$', '', s)
    return s


def _get_mes_ano_calendario(driver):
    """Lê o mês e ano atuais do calendário. Retorna (mes_int, ano_int) ou (None, None)."""
    try:
        titulo = driver.find_element(By.CSS_SELECTOR, ".fc-header-title").text.strip().lower()
        for nome, num in MESES_PT.items():
            if nome in titulo:
                anos = re.findall(r'\d{4}', titulo)
                if anos:
                    return (num, int(anos[0]))
    except Exception:
        pass
    return (None, None)


def _ir_para_mes(driver, ano, mes):
    """Navega o calendário para o mês/ano alvo usando prev ou next conforme necessário. """
    for _ in range(36): 
        mes_cal, ano_cal = _get_mes_ano_calendario(driver)
        if mes_cal is None:
            time.sleep(0.5)
            continue
        if (ano_cal, mes_cal) == (ano, mes):
            return True
        
        btn = ".fc-button-prev" if (ano_cal, mes_cal) > (ano, mes) else ".fc-button-next"
        try:
            driver.find_element(By.CSS_SELECTOR, btn).click()
            time.sleep(0.7)
        except Exception:
            break
    return False


def coletar_dados(ini, fim, salas, headless, log_fn, progresso_fn, set_total_fn):
    driver = criar_driver(headless)
    todos, urls_vistas, todos_links = [], set(), []

    try:
        log_fn("🔍 Fase 1/2 — Mapeando reservas...")
        for sala, cfg in salas.items():
            log_fn(f"   📋 {sala}...")
            driver.get(cfg["url"])
            try:
                WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, ".fc-header-title")))
            except Exception:
                log_fn(f"   ⚠️ Calendário não carregou para {sala}, pulando..."); continue
            time.sleep(1)

            mes = ini.replace(day=1)
            while mes <= fim:
                _ir_para_mes(driver, mes.year, mes.month)
                time.sleep(0.7)
                for ev in driver.find_elements(By.CSS_SELECTOR, "a.fc-event"):
                    href = ev.get_attribute("href")
                    if href and href not in urls_vistas:
                        urls_vistas.add(href)
                        todos_links.append((sala, href))
                mes = (mes.replace(month=mes.month+1)
                       if mes.month < 12 else mes.replace(year=mes.year+1, month=1))

        total = len(todos_links)
        log_fn(f"   ✅ {total} reservas únicas encontradas\n")
        log_fn("📥 Fase 2/2 — Coletando detalhes...")
        set_total_fn(total)

        vistos = set()

        for sala, href in todos_links:
            nome_curto = href.rstrip("/").split("/")[-1].replace("-"," ").title()[:35]
            progresso_fn(1, f"{sala} — {nome_curto}")
            alunos, datas, disc, codigo_disc, nome_disc, turno = _extrair_reserva(driver, href, ini, fim)
            base = slug_base(href)
            for d in datas:

                if alunos is not None:
                    chave = (sala, d, base, alunos)
                else:
                    chave = (sala, d, href)
                if chave not in vistos:
                    vistos.add(chave)
                    todos.append({"sala": sala, "disciplina": disc, "data": d, "alunos": alunos, "codigo_disc": codigo_disc, "nome_disc": nome_disc, "turno": turno})
            time.sleep(0.3)

    finally:
        driver.quit()

    return [r for r in todos if r["data"]]


def _extrair_reserva(driver, url, ini, fim):
    """Abre a URL da reserva em nova aba e extrai: disciplina, alunos, datas no período."""
    driver.execute_script("window.open('');")
    driver.switch_to.window(driver.window_handles[-1])
    try:
        driver.get(url)
        time.sleep(1.5)

        disc = url.rstrip("/").split("/")[-1].replace("-"," ").title()
        try:
            disc = driver.find_element(By.CSS_SELECTOR, "h1, .page-title").text.strip()
        except Exception:
            pass

        try:
            corpo = driver.find_element(By.CSS_SELECTOR, "article, .field--type-text-with-summary, .node__content")
            texto_corpo = corpo.text
        except Exception:
            texto_corpo = driver.find_element(By.CSS_SELECTOR, "body").text

        alunos = None
        m = re.search(r'[Qq]uantidade de [Aa]lunos[^\d]*(\d+)', texto_corpo)
        if m:
            alunos = int(m.group(1))

        codigo_disc = ""
        nome_disc   = ""
        m_cod  = re.search(r'C[oó]digo da Disciplina\s*/\s*Curso\s*[:\-]?\s*([^\n]+)', texto_corpo)
        m_nome = re.search(r'Nome da Disciplina\s*/\s*Curso\s*(?:\(opcional\))?\s*[:\-]?\s*([^\n]+)', texto_corpo)
        if m_cod:
            codigo_disc = m_cod.group(1).strip()
        if m_nome:
            nome_disc = m_nome.group(1).strip()

            nome_disc = re.sub(r'^[/\s]*Curso\s*\(opcional\)\s*[:\-]?\s*', '', nome_disc).strip()

        turno = ""
        m_hor = re.search(r'[Hh]or[aá]rios?[^\d]*(\d{2}:\d{2})\s*[àa]s?\s*(\d{2}:\d{2})', texto_corpo)
        if not m_hor:

            m_hor = re.search(r'das\s+(\d{2}:\d{2})\s+[àa]s\s+(\d{2}:\d{2})', texto_corpo)
        if not m_hor:

            m_hor = re.search(r'\d{2}/\d{2}/\d{4}\s*-\s*(\d{2}:\d{2})\s+até\s+(\d{2}:\d{2})', texto_corpo)
        if m_hor:
            turno = _calcular_turno(m_hor.group(1), m_hor.group(2))

        marcadores_fim = [
            "Nome do Solicitante", "Telefone", "E-mail", "Nome da Disciplina",
            "Código da Disciplina", "Codigo da Disciplina", "Software",
            "Aula Eventual", "Horários", "Horario", "Sala:", "Campus:",
            "Tipo de Reserva", "Dia(s) e horário", "Dia(s) e horario", "Observa",
        ]
        idx_inicio = texto_corpo.find("Data da Reserva")
        if idx_inicio >= 0:
            trecho = texto_corpo[idx_inicio:]
            idx_fim = len(trecho)
            for marcador in marcadores_fim:
                pos = trecho.find(marcador)
                if 0 < pos < idx_fim:
                    idx_fim = pos
            secao_data = trecho[:idx_fim]
        else:
            secao_data = texto_corpo 

        datas = []

        continua = re.search(
            r'(\d{2}/\d{2}/\d{4})\s*-\s*\d{2}:\d{2}\s+até\s+(\d{2}/\d{2}/\d{4})',
            secao_data
        )

        if continua:

            data_inicio = continua.group(1)
            try:
                dt = datetime.strptime(data_inicio, "%d/%m/%Y")
                if ini <= dt <= fim:
                    datas.append(data_inicio)
            except ValueError:
                pass
        else:

            achou_data = False
            for linha in secao_data.splitlines():
                ls = linha.strip()
                if not ls or ls.startswith("Repete"):
                    continue
                m_data = re.match(r'(\d{2}/\d{2}/\d{4})', ls)
                if m_data:
                    achou_data = True
                    d = m_data.group(1)
                    try:
                        dt = datetime.strptime(d, "%d/%m/%Y")
                        if ini <= dt <= fim:
                            datas.append(d)
                    except ValueError:
                        pass
                elif achou_data:
                    break

        return alunos, list(set(datas)), disc, codigo_disc, nome_disc, turno

    except Exception:
        return None, [], url.split("/")[-1], "", "", ""
    finally:
        driver.close()
        driver.switch_to.window(driver.window_handles[0])

def gerar_excel(dados, salas, campus_nome, ini, fim, destino):
    def fill(c): return PatternFill("solid", start_color=c, end_color=c)
    brd = Border(left=Side(style='thin',color='BFBFBF'), right=Side(style='thin',color='BFBFBF'),
                 top=Side(style='thin',color='BFBFBF'),  bottom=Side(style='thin',color='BFBFBF'))
    ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)
    esq = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    fh  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    fn  = Font(name="Arial", size=10)
    fb  = Font(name="Arial", size=10, bold=True)
    ft  = Font(name="Arial", size=10, bold=True, color="1F3864")

    tem_alunos = any(d['alunos'] is not None for d in dados)

    def stats(sala):
        r = [d for d in dados if d['sala']==sala]
        aulas  = len(set(d['data'] for d in r))
        alunos = sum(d['alunos'] for d in r if d['alunos']) if tem_alunos else None
        return aulas, alunos

    wb = Workbook()
    ws = wb.active; ws.title = "Relatório"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width=3; ws.column_dimensions['B'].width=46
    ws.column_dimensions['C'].width=18; ws.column_dimensions['D'].width=3

    ws.merge_cells('B2:C2')
    ws['B2'] = "Relatório de Utilização Vila Digital"
    ws['B2'].font=Font(name="Arial",size=16,bold=True,color="FFFFFF")
    ws['B2'].fill=fill("1F3864"); ws['B2'].alignment=ctr; ws['B2'].border=brd
    ws.row_dimensions[2].height=40

    ws.merge_cells('B3:C3')
    ws['B3'] = f"Campus {campus_nome} — {ini.strftime('%d/%m/%Y')} a {fim.strftime('%d/%m/%Y')}"
    ws['B3'].font=Font(name="Arial",size=11,bold=True,color="FFFFFF")
    ws['B3'].fill=fill("2E75B6"); ws['B3'].alignment=ctr; ws['B3'].border=brd
    ws.row_dimensions[3].height=28

    hr = 5
    if not tem_alunos:
        ws.merge_cells('B4:C4')
        ws['B4'] = "⚠ Este campus não registra quantidade de alunos nas reservas."
        ws['B4'].font=Font(name="Arial",size=9,italic=True,color="795548")
        ws['B4'].fill=fill("FFF8E1"); ws['B4'].alignment=ctr; ws['B4'].border=brd
        ws.row_dimensions[4].height=20; hr=6

    ws[f'B{hr}']="Descrição";  ws[f'B{hr}'].font=fh; ws[f'B{hr}'].fill=fill("1F3864"); ws[f'B{hr}'].alignment=ctr; ws[f'B{hr}'].border=brd
    ws[f'C{hr}']="Quantidade"; ws[f'C{hr}'].font=fh; ws[f'C{hr}'].fill=fill("1F3864"); ws[f'C{hr}'].alignment=ctr; ws[f'C{hr}'].border=brd
    ws.row_dimensions[hr].height=22

    total_aulas  = len(set(d['data'] for d in dados))
    total_alunos = sum(d['alunos'] for d in dados if d['alunos']) if tem_alunos else None
    total_pcs    = sum(v['computadores'] for v in salas.values())

    linhas = [("Aulas realizadas (datas únicas)", total_aulas, "F2F2F2")]
    if tem_alunos:
        linhas.append(("Alunos / aulas (total registrado)", total_alunos, "FFFFFF"))
    for sala in salas:
        a, al = stats(sala)
        cor = "F2F2F2" if len(linhas)%2==0 else "FFFFFF"
        linhas.append((f"{'Alunos' if tem_alunos else 'Aulas'} – Lab {sala}", al if tem_alunos else a, cor))
    for sala in salas:
        cor = "F2F2F2" if len(linhas)%2==0 else "FFFFFF"
        linhas.append((f"Laboratório {sala} – Computadores", salas[sala]['computadores'], cor))
    linhas.append(("Total de computadores nos laboratórios", total_pcs, "E2EFDA"))

    for i,(desc,val,cor) in enumerate(linhas):
        r=hr+1+i; ws.row_dimensions[r].height=22
        ws[f'B{r}']=desc; ws[f'B{r}'].font=fb if cor=="E2EFDA" else fn
        ws[f'B{r}'].fill=fill(cor); ws[f'B{r}'].alignment=esq; ws[f'B{r}'].border=brd
        ws[f'C{r}']=val if val is not None else "N/I"
        ws[f'C{r}'].font=ft if cor=="E2EFDA" else fn
        ws[f'C{r}'].fill=fill(cor); ws[f'C{r}'].alignment=ctr; ws[f'C{r}'].border=brd

    rn=hr+len(linhas)+2; ws.row_dimensions[rn].height=30
    ws.merge_cells(f'B{rn}:C{rn}')
    ws[f'B{rn}']=f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} | proplad.ufu.br/vila-digital"
    ws[f'B{rn}'].font=Font(name="Arial",size=8,italic=True,color="595959"); ws[f'B{rn}'].alignment=esq

    ws2=wb.create_sheet("Dados Detalhados"); ws2.sheet_view.showGridLines=False
    ws2.column_dimensions['A'].width=12; ws2.column_dimensions['B'].width=55
    ws2.column_dimensions['C'].width=15; ws2.column_dimensions['D'].width=12
    ws2.row_dimensions[1].height=28
    for col,h in zip(['A','B','C','D'],["Sala","Disciplina","Data","Alunos"]):
        ws2[f'{col}1']=h; ws2[f'{col}1'].font=fh
        ws2[f'{col}1'].fill=fill("1F3864"); ws2[f'{col}1'].alignment=ctr; ws2[f'{col}1'].border=brd
    paleta=["EBF3FB","EAF4E8","FEF9E7","F9EBEA","F4ECF7","E8F8F5","FDFEFE"]
    cores_sala={s:paleta[i%len(paleta)] for i,s in enumerate(salas.keys())}
    for i,r in enumerate(sorted(dados,key=lambda x:(x['sala'],x['data'] or ""))):
        row=i+2; cor=cores_sala.get(r['sala'],"FFFFFF")
        for col,val in zip(['A','B','C','D'],[r['sala'],r['disciplina'],r['data'] or "",r['alunos'] or "N/I"]):
            ws2[f'{col}{row}']=val; ws2[f'{col}{row}'].font=Font(name="Arial",size=9)
            ws2[f'{col}{row}'].fill=fill(cor); ws2[f'{col}{row}'].border=brd
            ws2[f'{col}{row}'].alignment=ctr if col!='B' else esq
        ws2.row_dimensions[row].height=15

    ws3=wb.create_sheet("Por Mês"); ws3.sheet_view.showGridLines=False
    salas_list=list(salas.keys()); cols_salas=[chr(66+i) for i in range(len(salas_list))]
    col_total=chr(66+len(salas_list))
    ws3.column_dimensions['A'].width=14
    for c in cols_salas+[col_total]: ws3.column_dimensions[c].width=14
    ws3.row_dimensions[1].height=28
    col_label="Aulas" if not tem_alunos else "Alunos"
    for col,h in zip(['A']+cols_salas+[col_total],["Mês"]+[f"{s}\n({col_label})" for s in salas_list]+["Total"]):
        ws3[f'{col}1']=h; ws3[f'{col}1'].font=fh
        ws3[f'{col}1'].fill=fill("1F3864"); ws3[f'{col}1'].alignment=ctr; ws3[f'{col}1'].border=brd

    meses_periodo = {}
    cur = ini.replace(day=1)
    while cur <= fim:
        meses_periodo[cur.month] = cur.strftime("%B").capitalize()
        cur = cur.replace(month=cur.month+1) if cur.month < 12 else cur.replace(year=cur.year+1, month=1)

    NOMES_MESES = {1:"Janeiro",2:"Fevereiro",3:"Março",4:"Abril",5:"Maio",6:"Junho",
                   7:"Julho",8:"Agosto",9:"Setembro",10:"Outubro",11:"Novembro",12:"Dezembro"}

    for j, mn in enumerate(sorted(meses_periodo.keys())):
        mnome = NOMES_MESES[mn]
        row=j+2; cor="F2F2F2" if j%2==0 else "FFFFFF"; ws3.row_dimensions[row].height=20
        ws3[f'A{row}']=mnome; ws3[f'A{row}'].font=Font(name="Arial",size=10,bold=True)
        ws3[f'A{row}'].fill=fill(cor); ws3[f'A{row}'].alignment=ctr; ws3[f'A{row}'].border=brd
        tot=0
        for col,sala in zip(cols_salas,salas_list):
            if tem_alunos:
                v=sum(d['alunos'] for d in dados if d['sala']==sala and d['alunos']
                      and d['data'] and datetime.strptime(d['data'],"%d/%m/%Y").month==mn)
            else:
                v=len(set(d['data'] for d in dados if d['sala']==sala and d['data']
                          and datetime.strptime(d['data'],"%d/%m/%Y").month==mn))
            tot+=v; ws3[f'{col}{row}']=v; ws3[f'{col}{row}'].font=fn
            ws3[f'{col}{row}'].fill=fill(cor); ws3[f'{col}{row}'].alignment=ctr; ws3[f'{col}{row}'].border=brd
        ws3[f'{col_total}{row}']=tot; ws3[f'{col_total}{row}'].font=Font(name="Arial",size=10,bold=True)
        ws3[f'{col_total}{row}'].fill=fill(cor); ws3[f'{col_total}{row}'].alignment=ctr; ws3[f'{col_total}{row}'].border=brd

    wb.save(destino)


def gerar_excel_disciplinas(dados, campus_nome, ini, fim, destino):
    """Gera Excel no formato de Informe, agrupado por sala, contendo:
    1. Tabela de Componentes Curriculares (Código e Denominação)
    2. Escala de utilização (Seg-Sab x Manhã/Tarde/Noite)
    """
    def fill(c): return PatternFill("solid", start_color=c, end_color=c)
    brd = Border(left=Side(style='thin',color='BFBFBF'), right=Side(style='thin',color='BFBFBF'),
                 top=Side(style='thin',color='BFBFBF'),  bottom=Side(style='thin',color='BFBFBF'))
    ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)
    esq = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    fh  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    fb  = Font(name="Arial", size=10, bold=True)
    fn  = Font(name="Arial", size=10)

    por_sala = {}
    for r in dados:
        sala = r["sala"]
        if sala not in por_sala:
            por_sala[sala] = {
                "componentes": {},
                "escala": {"Manhã": {i: set() for i in range(6)},
                           "Tarde": {i: set() for i in range(6)},
                           "Noite": {i: set() for i in range(6)}}
            }

        cod = r.get("codigo_disc", "").strip()
        nome = r.get("nome_disc", "").strip()

        if not cod and not nome:
            cod = r.get("disciplina", "Sem Código")

        if cod:
            por_sala[sala]["componentes"][cod] = nome

        turno = r.get("turno", "")
        data_str = r.get("data", "")

        if data_str and turno:
            try:
                dt = datetime.strptime(data_str, "%d/%m/%Y")
                wd = dt.weekday()
                if 0 <= wd <= 5:
                    for t in turno.split('/'):
                        if t in por_sala[sala]["escala"]:
                            por_sala[sala]["escala"][t][wd].add(cod)
            except ValueError:
                pass

    wb = Workbook()
    ws = wb.active
    ws.title = "Escala de Disciplinas"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18

    ws.merge_cells('B2:H2')
    ws['B2'] = "Escala de Disciplinas — Vila Digital UFU"
    ws['B2'].font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    ws['B2'].fill = fill("1F3864"); ws['B2'].alignment = ctr; ws['B2'].border = brd
    ws.row_dimensions[2].height = 40

    ws.merge_cells('B3:H3')
    ws['B3'] = f"Campus {campus_nome} — {ini.strftime('%d/%m/%Y')} a {fim.strftime('%d/%m/%Y')}"
    ws['B3'].font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    ws['B3'].fill = fill("2E75B6"); ws['B3'].alignment = ctr; ws['B3'].border = brd
    ws.row_dimensions[3].height = 28

    linha_atual = 5
    dias_semana = ["Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira", "Sexta-feira", "Sábado"]

    for sala, info in sorted(por_sala.items()):
        ws.merge_cells(f"B{linha_atual}:H{linha_atual}")
        ws[f"B{linha_atual}"] = f"Laboratório: {sala}"
        ws[f"B{linha_atual}"].font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        ws[f"B{linha_atual}"].fill = fill("1F3864")
        ws[f"B{linha_atual}"].alignment = ctr
        ws[f"B{linha_atual}"].border = brd
        ws.row_dimensions[linha_atual].height = 28
        linha_atual += 2

        ws.merge_cells(f"B{linha_atual}:H{linha_atual}")
        ws[f"B{linha_atual}"] = "Componente(s) curricular(es) de curso(s) de graduação realizado(s) cotidianamente no Laboratório"
        ws[f"B{linha_atual}"].font = fh
        ws[f"B{linha_atual}"].fill = fill("2E75B6")
        ws[f"B{linha_atual}"].border = brd
        ws[f"B{linha_atual}"].alignment = ctr
        ws.row_dimensions[linha_atual].height = 28
        linha_atual += 1

        ws[f"B{linha_atual}"] = "Código"
        ws[f"B{linha_atual}"].font = fh
        ws[f"B{linha_atual}"].fill = fill("1F3864")
        ws[f"B{linha_atual}"].border = brd
        ws[f"B{linha_atual}"].alignment = ctr
        ws.row_dimensions[linha_atual].height = 22

        ws.merge_cells(f"C{linha_atual}:H{linha_atual}")
        ws[f"C{linha_atual}"] = "Denominação"
        ws[f"C{linha_atual}"].font = fh
        ws[f"C{linha_atual}"].fill = fill("1F3864")
        ws[f"C{linha_atual}"].border = brd
        ws[f"C{linha_atual}"].alignment = ctr
        linha_atual += 1

        if not info["componentes"]:
            ws.merge_cells(f"B{linha_atual}:H{linha_atual}")
            ws[f"B{linha_atual}"] = "Nenhum componente registrado."
            ws[f"B{linha_atual}"].font = fn
            ws[f"B{linha_atual}"].fill = fill("F2F2F2")
            ws[f"B{linha_atual}"].border = brd
            ws[f"B{linha_atual}"].alignment = ctr
            ws.row_dimensions[linha_atual].height = 22
            linha_atual += 1
        else:
            for idx, (cod, denom) in enumerate(sorted(info["componentes"].items())):
                cor = "F2F2F2" if idx % 2 == 0 else "FFFFFF"
                ws[f"B{linha_atual}"] = cod
                ws[f"B{linha_atual}"].font = fn
                ws[f"B{linha_atual}"].fill = fill(cor)
                ws[f"B{linha_atual}"].border = brd
                ws[f"B{linha_atual}"].alignment = ctr
                ws.row_dimensions[linha_atual].height = 22

                ws.merge_cells(f"C{linha_atual}:H{linha_atual}")
                ws[f"C{linha_atual}"] = denom
                ws[f"C{linha_atual}"].font = fn
                ws[f"C{linha_atual}"].fill = fill(cor)
                ws[f"C{linha_atual}"].border = brd
                ws[f"C{linha_atual}"].alignment = esq
                linha_atual += 1

        linha_atual += 1

        ws.merge_cells(f"B{linha_atual}:H{linha_atual}")
        ws[f"B{linha_atual}"] = "Escala de utilização do Laboratório em período letivo\n(em cada turno, registre o código do componente curricular que cotidianamente utiliza o Laboratório)"
        ws[f"B{linha_atual}"].font = fh
        ws[f"B{linha_atual}"].fill = fill("2E75B6")
        ws[f"B{linha_atual}"].border = brd
        ws[f"B{linha_atual}"].alignment = ctr
        ws.row_dimensions[linha_atual].height = 35
        linha_atual += 1

        ws[f"B{linha_atual}"].fill = fill("1F3864")
        ws[f"B{linha_atual}"].border = brd
        ws.row_dimensions[linha_atual].height = 28
        for i, dia in enumerate(dias_semana):
            col = chr(67 + i)
            ws[f"{col}{linha_atual}"] = dia
            ws[f"{col}{linha_atual}"].font = fh
            ws[f"{col}{linha_atual}"].fill = fill("1F3864")
            ws[f"{col}{linha_atual}"].border = brd
            ws[f"{col}{linha_atual}"].alignment = ctr
        linha_atual += 1

        cores_turno = {"Manhã": "EBF3FB", "Tarde": "EAF4E8", "Noite": "FEF9E7"}
        for turno in ["Manhã", "Tarde", "Noite"]:
            cor_turno = cores_turno[turno]
            codigos_por_dia = [sorted(list(info["escala"][turno][i])) for i in range(6)]
            n_rows = max(max((len(c) for c in codigos_por_dia), default=0), 1)

            if n_rows > 1:
                ws.merge_cells(f"B{linha_atual}:B{linha_atual + n_rows - 1}")
            ws[f"B{linha_atual}"] = turno
            ws[f"B{linha_atual}"].font = fb
            ws[f"B{linha_atual}"].fill = fill(cor_turno)
            ws[f"B{linha_atual}"].border = brd
            ws[f"B{linha_atual}"].alignment = ctr

            for row_offset in range(n_rows):
                r = linha_atual + row_offset
                ws.row_dimensions[r].height = 20
                if row_offset > 0:
                    ws[f"B{r}"].fill = fill(cor_turno)
                    ws[f"B{r}"].border = brd
                for i in range(6):
                    col = chr(67 + i)
                    codigos = codigos_por_dia[i]
                    texto = codigos[row_offset] if row_offset < len(codigos) else ""
                    ws[f"{col}{r}"] = texto
                    ws[f"{col}{r}"].font = fn
                    ws[f"{col}{r}"].fill = fill(cor_turno)
                    ws[f"{col}{r}"].border = brd
                    ws[f"{col}{r}"].alignment = ctr

            linha_atual += n_rows

        linha_atual += 3

    rn = linha_atual
    ws.merge_cells(f'B{rn}:H{rn}')
    ws[f'B{rn}'] = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} | proplad.ufu.br/vila-digital"
    ws[f'B{rn}'].font = Font(name="Arial", size=8, italic=True, color="595959")
    ws[f'B{rn}'].alignment = esq

    wb.save(destino)

if __name__ == "__main__":
    app = App()
    app.mainloop()